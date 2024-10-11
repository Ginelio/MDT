# when [] is reference from text MDT.pdf
# comments denoted with ## come from the reviewer

# This file stays true to the original equations of the main paper studied

# 1. Import Libraries

import numpy as np
import pandas as pd
import gurobipy as gp
from pathlib import Path
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# 2. Define Parameters and Data [Parartima A & B]

def init_param():
    params = {
        # PV Parameters [Table 2]:
        'η_PV_REF': 0.181,
        'N_PV': 36,
        'P_PV_MAX': 0.225 * 36, # kW
        'A_PV': 1.244, # m^2
        'a': -0.0038, # 1/oC
        'NOCT': 45, # oC
        'T_REF': 25, # oC
        # Wind Turbine Parameters [Table 3]:
        'v_rated': 14, # m/s
        'P_rated': 3, # kW
        'v_cut_in': 3, # m/s
        'v_cut_out': 18, # m/s
        # Battery Parameters [Table 4]:
        'U_B': 12, # V
        'Q_B': 240, # Ah
        'N_B': 32,
        'C_B_IN': 400 * 32, # €
        'SOC_MAX': 0.90,
        'SOC_MIN': 0.60,
        'SOC_IN': 0.80,
        'P_B_MAX': 18, # kW
        'N_cycles': 1300,
        'η_B_CH': 0.9,
        'η_B_DIS': 0.9,
        'C_O_and_M_B': 0.01, # €
        # Electrolyzer Parameters [Table 5]:
        'P_EL_REF': 6, # kW
        'n_H2_EL_REF': 0.46, # mol/h 
        'P_EL_MIN': 1.5, # kW
        'P_EL_MAX': 6.2, # kW
        'L_EL': 30000, # h
        'C_EL_IN': 75000, # €
        'C_O_and_M_EL': 0.001,  # €/h
        'η_EL': 0.7,
        # Hydrogen Storage Parameters [Table 5]:
        'p_H2_MAX': 13.8, # bar
        'p_H2_MIN': 2, # bar
        'p_H2_IN': 10, # bar
        'T_H2':  313, # K
        'V_H2': 4, # m^3
        # Fuel Cells Parameters [Table 5]:
        'P_NOM_FC': 5, # kW
        'n_H2_FC_REF': 0.17, # mol/h 
        'P_FC_MIN': 0.5, # kW
        'P_FC_MAX': 6, # kW
        'C_FC_IN': 28000, # €
        'L_FC': 30000, # h
        'C_O_and_M_FC': 0.00001,  # €/h
        'η_FC': 0.6,
        # General Parameters:
        'LHV_H2': 240, # kJ/mol [Equation 9]
        'R': 0.08314,  # m^3*bar/(mol*K) [Equation 8]
        'C_UN': 0.5,  # €/kWh [Equation 14]
        'C_EX': 0.5,   # €/kWh [Equation 14]
        'M': 1000,  # Big M value
    }
    return params

# Initialize the parameters
parameters = init_param()

# Load the data
file_path = Path('data_fall.csv')
try:
    df = pd.read_csv(file_path)
except FileNotFoundError:
    raise FileNotFoundError("The data_fall.csv file does not exist.")

# Extracting Columns:
T_AMB = df['T_AMB']
G = df['G']
v = df['v']
P_LD = df['P_LD']

# Load workbook and sheet
wb = load_workbook('20240910_EL-DAM_Results_EN_v01.xlsx', data_only=True)
ws = wb['EL-DAM_Results']  # Replace with the correct sheet name

# Extract values from cells J411 to J434 (for example) and convert to €/kWh
mcp_sell = [ws[f'J{row}'].value / 1000 for row in range(411, 435)]
mcp_buy = [ws[f'J{row}'].value / 1000 for row in range(155, 179)]

# Define Functions

# PV Power Output
def P_PV(T_AMB, G):
    # [Equation 2]
    η_PV = parameters['η_PV_REF'] * (1 - parameters['a'] * (T_AMB + (G * ((parameters['NOCT'] - 20) / 800)) - parameters['T_REF']))
    # [Equation 1]
    P_PV = (G * parameters['A_PV'] * parameters['N_PV'] * η_PV) / 1000 # kW
    # P_PV must be within the limits of 0 and P_PV_MAX
    P_PV = max(0, min(P_PV, parameters['P_PV_MAX']))
    return P_PV

# Wind Turbine Power Output [Equation 3]
def P_WT(v):
    if v < parameters['v_cut_in'] or v >= parameters['v_cut_out']:
        return 0 # kW
    elif parameters['v_cut_in'] <= v < parameters['v_rated']:
        return ((v - parameters['v_cut_in']) / (parameters['v_rated'] - parameters['v_cut_in'])) * parameters['P_rated'] # kW
    elif parameters['v_rated'] <= v < parameters['v_cut_out']:
        return parameters['P_rated'] # kW
    else:
        return 0 # kW 

# Battery charging cost [Equation 10]
def C_B_CH(P_B_CH, Y_B_CH, parameters):
    C_B_CH = (((parameters['C_B_IN'] * P_B_CH) / (parameters['N_B'] * parameters['U_B'] * parameters['Q_B'] * parameters['N_cycles'])) + Y_B_CH * parameters['C_O_and_M_B']) / (parameters['η_B_CH'] * parameters['η_B_DIS'])
    return C_B_CH

# Battery discharging cost [Equation 11]
def C_B_DIS(P_B_DIS, Y_B_DIS, parameters):
    C_B_DIS = ((parameters['C_B_IN'] * P_B_DIS) / (parameters['N_B'] * parameters['U_B'] * parameters['Q_B'] * parameters['N_cycles'] * parameters['η_B_DIS'])) + Y_B_DIS * parameters['C_O_and_M_B']
    return C_B_DIS

# Battery charging cost [Equation 10]
def C_B_CH(P_B_CH, Y_B_CH, parameters):
    C_B_CH = (((parameters['C_B_IN'] * P_B_CH) / (parameters['N_B'] * parameters['U_B'] * parameters['Q_B'] * parameters['N_cycles'])) + Y_B_CH * parameters['C_O_and_M_B']) / (parameters['η_B_CH'] * parameters['η_B_DIS'])
    return C_B_CH

# Fuel cell cost - Hydrogen consumption cost [Equation 14]
def C_H2_CH(P_EL, Y_EL, Y_FC, parameters):
    C_H2_CH = (((parameters['C_EL_IN'] / parameters['L_EL']) + (Y_EL * parameters['C_O_and_M_EL'])) / (parameters['η_EL'] * parameters['η_FC']) + (parameters['C_FC_IN'] / parameters['L_FC']) + (Y_FC * parameters['C_O_and_M_FC']) * (P_EL / parameters['P_NOM_FC']))
    return C_H2_CH

# Fuel cell cost - Hydrogen consumption cost [Equation 14]
def C_FC(Y_FC, parameters):
    C_FC = ((parameters['C_FC_IN'] / parameters['L_FC']) + Y_FC * parameters['C_O_and_M_FC'])
    return C_FC

# Excess electricity cost
def C_EX(P_EX, mcp_sell, parameters):
    C_EX = - mcp_sell * P_EX
    return C_EX

# Undelivered electricity cost
def C_UN(P_UN, mcp_buy, parameters):
    C_UN = mcp_buy * P_UN
    return C_UN

# Initialize a dictionary to store the cost components and other values for each time period
results = {
    'Time Period': [],
    'P_PV': [],
    'P_WT': [],
    'P_LD': [],
    'P_B_CH': [],
    'P_B_DIS': [],
    'Y_B_CH': [],
    'Y_B_DIS': [],
    'SOC': [],
    'P_EL': [],
    'Y_EL': [],
    'n_H2_EL': [],
    'p_H2': [],
    'P_FC': [],
    'Y_FC': [],
    'n_H2_FC': [],
    'P_EX': [],
    'P_UN': [],
    'z': [],
    'C_B_CH': [],
    'C_B_DIS': [],
    'C_H2_CH': [],
    'C_FC': [],
    'C_EX': [],
    'C_UN': [],
    'mcp_sell': [],
    'mcp_buy': [],
    'Total Cost': []
}

def objective_function(P_B_CH, P_B_DIS, Y_B_CH, Y_B_DIS, Y_EL, Y_FC, P_EX, P_UN, mcp_sell, mcp_buy):
    obj = gp.LinExpr()
    for t in range(num_periods):
        C_B_CH_val = C_B_CH(P_B_CH[t], Y_B_CH[t], parameters)
        C_B_DIS_val = C_B_DIS(P_B_DIS[t], Y_B_DIS[t], parameters)
        C_H2_CH_val = C_H2_CH(P_EL[t], Y_EL[t], Y_FC[t], parameters)
        C_FC_val = C_FC(Y_FC[t], parameters)
        C_EX_val = C_EX(P_EX[t], mcp_sell[t], parameters)
        C_UN_val = C_UN(P_UN[t], mcp_buy[t], parameters)
        total_cost = C_B_CH_val + C_B_DIS_val + C_H2_CH_val + C_FC_val + C_EX_val + C_UN_val
        
        obj += total_cost
    return obj

model = gp.Model("Microgrid Optimization")
num_periods = len(df)

# Call Parameters for PV Power Output and Wind Turbine Power Output
P_PV_values = [P_PV(T_AMB[t], G[t]) for t in range(num_periods)]
P_WT_values = [P_WT(v[t]) for t in range(num_periods)]
P_LD_values = [P_LD[t] for t in range(num_periods)]

# 4. Define Decision Variables 

# Battery System
P_B_CH = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, ub=parameters['P_B_MAX'], name="P_B_CH") # [Equation 19]
P_B_DIS = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, ub=parameters['P_B_MAX'], name="P_B_DIS") # [Equation 20]
SOC = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=parameters['SOC_MIN'], ub=parameters['SOC_MAX'], name="SOC") # [Equaton 26] 
Y_B_CH = model.addVars(num_periods, vtype=gp.GRB.BINARY, name="Y_B_CH")
Y_B_DIS = model.addVars(num_periods, vtype=gp.GRB.BINARY, name="Y_B_DIS")
# Electrolyzer
P_EL = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, ub=parameters['P_EL_MAX'], name="P_EL") # [Equation 21]
n_H2_EL = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, ub=parameters['n_H2_EL_REF'], name="n_H2_EL")
Y_EL = model.addVars(num_periods, vtype=gp.GRB.BINARY, name="Y_EL")
# Hydrogen Storage
p_H2 = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=parameters['p_H2_MIN'], ub=parameters['p_H2_MAX'], name="p_H2") # [Equation 27]
# Fuel Cell
P_FC = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, ub=parameters['P_FC_MAX'], name="P_FC") # [Equation 22]
n_H2_FC = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, ub=parameters['n_H2_FC_REF'], name="n_H2_FC")
Y_FC = model.addVars(num_periods, vtype=gp.GRB.BINARY, name="Y_FC")
#  Power Balance Variables
P_UN = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, name="P_UN") # [Equation 24]
P_EX = model.addVars(num_periods, vtype=gp.GRB.CONTINUOUS, lb=0, name="P_EX") # [Equation 25]
z = model.addVars(num_periods, vtype=gp.GRB.BINARY, name="z")

# Objective function
model.setObjective(objective_function(P_B_CH, P_B_DIS, Y_B_CH, Y_B_DIS, Y_EL, Y_FC, P_EX, P_UN, mcp_sell, mcp_buy), gp.GRB.MINIMIZE)

# 7. Adding Constraints

# SOC constraint [Equation 4]
model.addConstr(SOC[0] == parameters['SOC_IN'] + (P_B_CH[0] * parameters['η_B_CH'] - P_B_DIS[0] / parameters['η_B_DIS']) * 1000 /(parameters['N_B'] * parameters['U_B'] * parameters['Q_B']), name="SOC_Balance_0}")
for t in range(1, num_periods):
    model.addConstr(SOC[t] == SOC[t-1] + (P_B_CH[t] * parameters['η_B_CH'] - P_B_DIS[t] / parameters['η_B_DIS']) * 1000 /(parameters['N_B'] * parameters['U_B'] * parameters['Q_B']), name=f"SOC_Balance_{t}") # [Equation 4]

# Hydrogen Pressure [Equation 8]:
model.addConstr(p_H2[0] == parameters['p_H2_IN'] + ((parameters['R'] * parameters['T_H2']) / (parameters['V_H2'])) * (n_H2_EL[0] - n_H2_FC[0]), name="Hydrogen_Pressure_0")
for t in range(1, num_periods):
    model.addConstr(p_H2[t] == p_H2[t-1] + ((parameters['R'] * parameters['T_H2']) / parameters['V_H2']) * (n_H2_EL[t] - n_H2_FC[t]), name=f"Hydrogen_Pressure_{t}")  # [Equation 8]  

# Power balance
for t in range(0, num_periods):
    model.addConstr(P_PV_values[t] + P_WT_values[t] - P_LD_values[t] + P_B_DIS[t] + P_FC[t] + P_UN[t] - P_B_CH[t] - P_EL[t] - P_EX[t] == 0, name=f"Power_Balance_{t}") # [Equation 18]

# Battery charge
for t in range(0, num_periods):
    model.addConstr(P_B_CH[t] - (parameters['P_B_MAX'] * Y_B_CH[t]) <= 0, name=f"Battery_charge_{t}") # Equation 26

# Battery discharge
for t in range(0, num_periods):
    model.addConstr(P_B_DIS[t] - (parameters['P_B_MAX'] * Y_B_DIS[t]) <= 0, name=f"Battery_discharge_{t}") # Equation 27

# Battery charging or discharging
for t in range(0, num_periods):
    model.addConstr(Y_B_CH[t] + Y_B_DIS[t] <= 1, name=f"Battery_charging_discharging_{t}")

# Electrolyzer max
for t in range(0, num_periods):
    model.addConstr(P_EL[t] - parameters['P_EL_MAX'] * Y_EL[t] <= 0, name=f"EL_max_{t}") # Equation 29

# Electrolyzer min
for t in range(0, num_periods):
    model.addConstr(P_EL[t] - parameters['P_EL_MIN'] * Y_EL[t] >= 0, name=f"EL_min_{t}")

# Electrolyzer activ
for t in range(0, num_periods):
    model.addConstr(n_H2_EL[t] - parameters['n_H2_EL_REF'] * Y_EL[t] <= 0, name=f"EL_on_{t}")

# Electrolyzer production
for t in range(0, num_periods):
    model.addConstr(n_H2_EL[t] == (parameters['η_EL'] * P_EL[t]) / parameters['LHV_H2'], name=f"EL_production_{t}")  # Produced Hydrogen Flow [Equation 7]

# Fuel cell max
for t in range(0, num_periods):
    model.addConstr(P_FC[t] - parameters['P_FC_MAX'] * Y_FC[t] <= 0, name=f"FC_max_{t}") # Equation 31

# Fuel cell min
for t in range(0, num_periods):
    model.addConstr(P_FC[t] - parameters['P_FC_MIN'] * Y_FC[t] >= 0, name=f"FC_min_{t}")

# Fuel cell activ
for t in range(0, num_periods):
    model.addConstr(n_H2_FC[t] - parameters['n_H2_FC_REF'] * Y_FC[t] <= 0, name=f"FC_activ_{t}")

# Fuel cell consumption
for t in range(0, num_periods):
    model.addConstr(n_H2_FC[t] == (P_FC[t]) / (parameters['η_FC'] * parameters['LHV_H2']), name=f"FC_consumption_{t}")  # Equation 9

# H2 charging discharging
for t in range(0, num_periods):
    model.addConstr(Y_FC[t] + Y_EL[t] <= 1, name=f"H2_charging_discharging_{t}")

# Undelivered power
for t in range(0, num_periods):
    model.addConstr(P_UN[t] - parameters['M'] * z[t] <= 0, name=f"Undelivered_power_{t}") # Equation 32

# Excess power
for t in range(0, num_periods):
    model.addConstr(P_EX[t] - parameters['M'] * (1 - z[t]) <= 0, name=f"Excess_power_{t}") # Equation 33

# State of Charge at the end of the day
model.addConstr(SOC[num_periods-1] == parameters['SOC_IN'], name="SOC_End")

# Hydrogen Storage at the end of the day
model.addConstr(p_H2[num_periods-1] == parameters['p_H2_IN'], name="Hydrogen_Storage_End")

# Optimize the model
model.optimize()

# model.display()

# 6. Output the results

# Check if the optimization was successful
if model.status == gp.GRB.OPTIMAL:
    print("Optimal solution found.", model.objVal)
elif model.status == gp.GRB.INFEASIBLE:
    print("Optimal Solution Not Found - Model is Infeasible")
elif model.status == gp.GRB.UNBOUNDED:
    print("Optimal Solution Not Found - Model is Unbounded")
elif model.status == gp.GRB.INF_OR_UNBD:
    print("Optimal Solution Not Found - Model is either Infeasible or Unbounded")
elif model.status == gp.GRB.TIME_LIMIT:
    print("Optimal Solution Not Found - Time limit reached")

# Extract the values from the Gurobi variables and append to the dictionary
for t in range(num_periods):
        results['Time Period'].append(t)
        results['P_PV'].append(round(P_PV_values[t],1))
        results['P_WT'].append(round(P_WT_values[t],1))
        results['P_LD'].append(round(P_LD_values[t],1))
        results['P_B_CH'].append(round(P_B_CH[t].X,1))
        results['P_B_DIS'].append(round(P_B_DIS[t].X,1))
        results['Y_B_CH'].append(Y_B_CH[t].X)
        results['Y_B_DIS'].append(Y_B_DIS[t].X)
        results['SOC'].append(round(SOC[t].X,2))
        results['P_EL'].append(round(P_EL[t].X,1))
        results['Y_EL'].append(Y_EL[t].X)
        results['n_H2_EL'].append(round(n_H2_EL[t].X,2))
        results['p_H2'].append(round(p_H2[t].X,1))
        results['P_FC'].append(round(P_FC[t].X,1))
        results['Y_FC'].append(Y_FC[t].X)
        results['n_H2_FC'].append(round(n_H2_FC[t].X,2))
        results['P_EX'].append(round(P_EX[t].X,1))
        results['P_UN'].append(round(P_UN[t].X,1))
        results['z'].append(z[t].X)
        results['C_B_CH'].append(C_B_CH(P_B_CH[t], Y_B_CH[t], parameters).getValue())
        results['C_B_DIS'].append(C_B_DIS(P_B_DIS[t], Y_B_DIS[t], parameters).getValue())
        results['C_H2_CH'].append(C_H2_CH(P_EL[t], Y_EL[t], Y_FC[t], parameters).getValue())
        results['C_FC'].append(C_FC(Y_FC[t], parameters).getValue())
        results['C_EX'].append(C_EX(P_EX[t], mcp_sell[t], parameters).getValue())
        results['C_UN'].append(C_UN(P_UN[t], mcp_buy[t], parameters).getValue())
        results['mcp_sell'].append(round(mcp_sell[t],3))
        results['mcp_buy'].append(round(mcp_buy[t],3))
        results['Total Cost'].append((C_B_CH(P_B_CH[t], Y_B_CH[t], parameters) + 
                                      C_B_DIS(P_B_DIS[t], Y_B_DIS[t], parameters) + 
                                      C_H2_CH(P_EL[t], Y_EL[t], Y_FC[t], parameters) + 
                                      C_FC(Y_FC[t], parameters) + 
                                      C_EX(P_EX[t], mcp_sell[t], parameters) + 
                                      C_UN(P_UN[t], mcp_buy[t], parameters)).getValue())

# Convert the results dictionary to a DataFrame
df_results = pd.DataFrame(results)

# Define the path to the Excel file
excel_file_path = 'C:/Users/gkara/Documents/Master/Master Diploma Thesis/DT/Data/results.xlsx'

# Write the DataFrame to an Excel file
df_results.to_excel(excel_file_path, index=False)

print(f"Results have been written to {excel_file_path}")

# Check if the model is infeasible
if model.status == gp.GRB.INFEASIBLE:
    print("Model is infeasible. Computing IIS...")
    model.computeIIS()
    model.write("model.ilp")

# Calculate CO₂ emissions

# Emission factor for natural gas in kg CO₂ per kWh
EF_natural_gas = 0.27

# Assuming 'results' is a DataFrame with 'P_ABS' column representing power absorbed from the grid, here is P_UN, which is the power that is absorbed from the grid
# Read your data if necessary
results_df = pd.read_excel(excel_file_path)

# Calculate hourly CO₂ emissions
results_df['CO2_Emissions'] = results_df['P_UN'] * EF_natural_gas

# Save results to a file
results_df.to_csv("results_with_CO2_emissions.csv", index=False)
print("Hourly CO₂ emissions calculated and saved to results_with_CO2_emissions.csv")